"""Upload page for dashboard-only estimator refresh."""

from __future__ import annotations

import re
import shutil
from pathlib import Path

import pandas as pd
import streamlit as st
from streamlit_echarts import st_echarts

from src.app.components.echarts_utils import build_stacked_bar_chart_options

ROOT = Path(__file__).resolve().parents[3]
UPLOAD_DIR = ROOT / "data" / "uploads"
RAW_DIR = ROOT / "data" / "raw"

DASHBOARD_WORKBOOK_NAME = "20260422_Data for Dashboard.xlsx"
SUPPORTED_EXTENSIONS = {".xlsx"}


def _ensure_dirs() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)


def _save_uploaded_file(uploaded_file) -> Path:
    _ensure_dirs()
    dest = UPLOAD_DIR / uploaded_file.name
    dest.write_bytes(uploaded_file.getvalue())
    return dest


def _extract_campaign_year(value: object) -> int | None:
    match = re.search(r"(20\d{2})", str(value or ""))
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _read_sheets_for_validation(file_path: Path) -> dict[str, list[list[str]]] | None:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        sheets_data: dict[str, list[list[str]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 50:
                    break
                rows.append([str(cell).strip() if cell is not None else "" for cell in row])
            sheets_data[sheet_name] = rows
        wb.close()
        return sheets_data
    except Exception:
        return None


def _preview_excel_sheet(file_path: Path, sheet_name: str) -> None:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return

        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6:
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()

        if rows:
            st.markdown(f"**{sheet_name}** — Preview (first {len(rows)} rows)")
            st.dataframe(rows, width="stretch", hide_index=True)
    except Exception as exc:
        st.warning(f"Could not preview `{sheet_name}`: {exc}")


def _copy_to_raw_dashboard(file_path: Path) -> Path:
    _ensure_dirs()
    for existing in RAW_DIR.glob("*"):
        if existing.is_file() and existing.suffix.lower() in SUPPORTED_EXTENSIONS:
            existing.unlink()
    dest = RAW_DIR / DASHBOARD_WORKBOOK_NAME
    shutil.copy2(file_path, dest)
    return dest


def _read_dashboard_frame(file_path: Path, sheet_name: str, *, usecols: list[str]) -> pd.DataFrame:
    try:
        frame = pd.read_excel(file_path, sheet_name=sheet_name, header=1, usecols=usecols, engine="openpyxl").fillna("")
    except Exception:
        return pd.DataFrame()

    for column in ["Actual Cost USD", "Actual depth, ft MD", "Drilling Duration, days equivalent"]:
        if column in frame.columns:
            frame[column] = pd.to_numeric(
                frame[column].astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            )
    return frame


def _load_workbook_from_session(existing_dashboard: Path) -> Path | None:
    for key in ("upload_staged_path", "upload_raw_dashboard_path"):
        stored = st.session_state.get(key)
        if stored:
            candidate = Path(stored)
            if candidate.exists():
                return candidate
    if existing_dashboard.exists():
        return existing_dashboard
    return None


def _render_loaded_data_dashboard(file_path: Path) -> None:
    structured = _read_dashboard_frame(
        file_path,
        "Structured.Cost",
        usecols=["Asset", "Campaign", "Level 4", "Well", "Actual Cost USD"],
    )
    campaign_meta = _read_dashboard_frame(
        file_path,
        "General.Camp.Data",
        usecols=["Asset", "Campaign", "Well Name\nActual", "Actual depth, ft MD", "Drilling Duration, days equivalent"],
    )
    if structured.empty or campaign_meta.empty:
        st.info("Loaded data dashboard will appear once the workbook passes validation and the required sheets can be read.")
        return

    for frame in (structured, campaign_meta):
        for column in ("Asset", "Campaign"):
            frame[column] = frame[column].astype(str).str.strip()

    structured["Level 4"] = structured["Level 4"].astype(str).str.strip()
    structured["Well"] = structured["Well"].astype(str).str.strip()
    campaign_meta["Well Name\nActual"] = campaign_meta["Well Name\nActual"].astype(str).str.strip()

    structured["campaign_year"] = structured["Campaign"].map(_extract_campaign_year)
    campaign_meta["campaign_year"] = campaign_meta["Campaign"].map(_extract_campaign_year)

    valid_wells = campaign_meta[campaign_meta["Well Name\nActual"].replace("", pd.NA).notna()].copy()
    campaign_register = (
        valid_wells.groupby(["Asset", "Campaign", "campaign_year"], dropna=False)["Well Name\nActual"]
        .nunique()
        .reset_index(name="well_count")
    )
    campaign_register["campaign_year"] = campaign_register["campaign_year"].fillna(0).astype(int)

    cost_register = (
        structured.groupby(["Asset", "Campaign", "campaign_year"], dropna=False)["Actual Cost USD"]
        .sum()
        .reset_index()
    )
    cost_register["campaign_year"] = cost_register["campaign_year"].fillna(0).astype(int)
    cost_register["total_cost_mmusd"] = cost_register["Actual Cost USD"] / 1_000_000.0

    top_l4_register = (
        structured.groupby(["Level 4", "Asset", "campaign_year"], dropna=False)["Actual Cost USD"]
        .sum()
        .reset_index()
    )
    top_l4_register["campaign_year"] = top_l4_register["campaign_year"].fillna(0).astype(int)
    top_l4_register["field_year"] = top_l4_register.apply(
        lambda row: f"{row['Asset']} · {row['campaign_year']}" if row["campaign_year"] else str(row["Asset"]),
        axis=1,
    )
    top_l4_register["total_cost_mmusd"] = top_l4_register["Actual Cost USD"] / 1_000_000.0

    fields_loaded = [field for field in sorted(valid_wells["Asset"].dropna().unique().tolist()) if field]
    campaign_list = [campaign for campaign in sorted(cost_register["Campaign"].dropna().unique().tolist()) if campaign]
    total_cost_mmusd = float(structured["Actual Cost USD"].sum()) / 1_000_000.0
    total_wells = int(valid_wells["Well Name\nActual"].nunique())
    total_campaigns = int(cost_register["Campaign"].replace("", pd.NA).dropna().nunique())
    total_l4 = int(structured["Level 4"].replace("", pd.NA).dropna().nunique())
    year_values = [year for year in sorted(cost_register["campaign_year"].unique().tolist()) if year]

    st.markdown("**Loaded data dashboard**")
    st.caption(f"Fields loaded: {', '.join(fields_loaded) if fields_loaded else 'n/a'}")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Fields loaded", f"{len(fields_loaded):,}")
    c2.metric("Campaigns", f"{total_campaigns:,}")
    c3.metric("Wells", f"{total_wells:,}")
    c4.metric("Total cost (MMUSD)", f"{total_cost_mmusd:,.2f}")
    c5.metric("Lv.4 families", f"{total_l4:,}")
    c6.metric("Year span", f"{min(year_values)}-{max(year_values)}" if year_values else "n/a")

    years = [year for year in sorted(campaign_register["campaign_year"].unique().tolist()) if year]
    field_categories = fields_loaded

    well_stack = []
    cost_stack = []
    palette = [
        "#0f766e",
        "#2563eb",
        "#7c3aed",
        "#ea580c",
        "#dc2626",
        "#16a34a",
        "#0891b2",
        "#9333ea",
    ]
    for idx, year in enumerate(years):
        well_values = []
        cost_values = []
        for field in field_categories:
            field_year_wells = campaign_register[
                (campaign_register["Asset"] == field) & (campaign_register["campaign_year"] == year)
            ]["well_count"].sum()
            field_year_cost = cost_register[
                (cost_register["Asset"] == field) & (cost_register["campaign_year"] == year)
            ]["total_cost_mmusd"].sum()
            well_values.append(int(field_year_wells))
            cost_values.append(round(float(field_year_cost), 3))
        well_stack.append({"name": str(year), "data": well_values, "color": palette[idx % len(palette)]})
        cost_stack.append({"name": str(year), "data": cost_values, "color": palette[idx % len(palette)]})

    stacked_cols = st.columns(2)
    with stacked_cols[0]:
        st_echarts(
            build_stacked_bar_chart_options(
                "Wells by field stacked by campaign year",
                field_categories,
                well_stack,
                unit="wells",
                horizontal=True,
                integer_labels=True,
                show_labels=True,
            ),
            height="380px",
            key="upload_wells_by_field_year",
        )
    with stacked_cols[1]:
        st_echarts(
            build_stacked_bar_chart_options(
                "Total cost by field stacked by campaign year",
                field_categories,
                cost_stack,
                unit="MMUSD",
                horizontal=True,
                integer_labels=False,
                show_labels=False,
            ),
            height="380px",
            key="upload_cost_by_field_year",
        )

    top_field_years = (
        top_l4_register.groupby("field_year", dropna=False)["total_cost_mmusd"].sum().sort_values(ascending=False).head(8).index.tolist()
    )
    top_l4_labels = (
        top_l4_register.groupby("Level 4", dropna=False)["total_cost_mmusd"].sum().sort_values(ascending=False).head(10).index.tolist()
    )
    l4_series = []
    for idx, field_year in enumerate(top_field_years):
        series_values = []
        for level_4 in top_l4_labels:
            total = top_l4_register[
                (top_l4_register["field_year"] == field_year) & (top_l4_register["Level 4"] == level_4)
            ]["total_cost_mmusd"].sum()
            series_values.append(round(float(total), 3))
        l4_series.append({"name": field_year, "data": series_values, "color": palette[idx % len(palette)]})

    st_echarts(
        build_stacked_bar_chart_options(
            "Top Lv.4 cost families stacked by field-campaign year",
            top_l4_labels,
            l4_series,
            unit="MMUSD",
            horizontal=True,
            integer_labels=False,
            show_labels=False,
        ),
        height="460px",
        key="upload_level4_by_field_year",
    )

    st.caption(
        f"Campaigns in this session: {', '.join(campaign_list) if campaign_list else 'n/a'} | "
        "The first row now shows the field mix directly, while the Lv.4 row shows where that mix is landing inside the cost structure."
    )

    with st.expander("Loaded campaign register", expanded=False):
        merged = campaign_register.merge(
            cost_register[["Asset", "Campaign", "campaign_year", "total_cost_mmusd"]],
            on=["Asset", "Campaign", "campaign_year"],
            how="left",
        ).rename(
            columns={
                "Asset": "field",
                "Campaign": "campaign",
                "campaign_year": "campaign_year",
                "well_count": "well_count",
                "total_cost_mmusd": "total_cost_mmusd",
            }
        )
        st.dataframe(
            merged.sort_values(["campaign_year", "field", "campaign"], ascending=[True, True, True]),
            width="stretch",
            hide_index=True,
        )


def render_upload_page() -> None:
    st.markdown("# DATA UPLOAD")
    st.markdown(
        "Upload the dashboard workbook used by the estimator refresh pipeline. "
        "Only one file is required: **20260422_Data for Dashboard.xlsx** (or same schema with a different filename)."
    )

    _ensure_dirs()
    existing_dashboard = RAW_DIR / DASHBOARD_WORKBOOK_NAME
    if existing_dashboard.exists():
        size_kb = existing_dashboard.stat().st_size / 1024
        st.info(f"Current raw workbook: `{existing_dashboard.name}` ({size_kb:.1f} KB)")

    st.divider()

    uploaded_file = st.file_uploader(
        "Upload dashboard workbook (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        key="dashboard_workbook_uploader",
        help="This upload replaces the existing raw workbook used by the estimator pipeline.",
    )

    staged_path: Path | None = None
    if uploaded_file is not None:
        staged_path = _save_uploaded_file(uploaded_file)
        st.session_state["upload_staged_path"] = str(staged_path)
    else:
        staged_path = _load_workbook_from_session(existing_dashboard)

    if staged_path is None or not staged_path.exists():
        st.info("Upload one dashboard workbook to begin.")
        return

    if uploaded_file is None:
        st.caption(f"Using cached workbook for this session: `{staged_path.name}`")

    from src.app.components.data_validation import validate_excel_file

    sheets_data = _read_sheets_for_validation(staged_path)
    if sheets_data is None:
        st.error("Could not read workbook. Please upload a valid, unencrypted .xlsx file.")
        return

    validation = validate_excel_file(staged_path, sheets_data)

    if validation.pipeline_ready:
        st.success(f"Workbook validated: **{validation.file_name}**")
    else:
        st.error("Workbook did not pass required dashboard contract checks.")

    for warning in validation.warnings:
        st.warning(warning)

    with st.expander("Validation details", expanded=not validation.pipeline_ready):
        for sr in validation.sheet_results:
            if sr.headers_valid:
                st.markdown(f"  ✓ `{sr.sheet_name}` — {sr.row_count} rows")
            elif sr.found:
                st.markdown(f"  ✗ `{sr.sheet_name}` — Missing headers: {', '.join(sr.missing_headers)}")
            else:
                st.markdown(f"  — `{sr.sheet_name}` — Sheet not found")

    if st.button("CONFIRM & LOAD DATA", type="primary", disabled=not validation.pipeline_ready):
        raw_path = _copy_to_raw_dashboard(staged_path)
        st.session_state["upload_confirmed"] = True
        st.session_state["validation_results"] = [validation]
        st.session_state["upload_raw_dashboard_path"] = str(raw_path)
        st.success(
            f"Loaded dashboard workbook as `{raw_path.name}`. "
            "Proceed to **Build Artifacts** to refresh estimator datasets."
        )

    if validation.pipeline_ready:
        _render_loaded_data_dashboard(staged_path)

    with st.expander("Data previews", expanded=False):
        st.caption("Sheet previews are collapsed so the dashboard stays front and center.")
        for sheet_name in ["Structured.Cost", "General.Camp.Data", "NPT.Data", "Check.Total"]:
            _preview_excel_sheet(staged_path, sheet_name)
